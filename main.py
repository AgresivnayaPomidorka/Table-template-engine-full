import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl

def delete_selected_item():
    selected_item = treeview.focus()
    if selected_item:
        result = messagebox.askyesno("Подтверждение", "Вы уверены, что хотите удалить выбранный элемент?")
        if result == tk.YES:
            # Удаление выбранного элемента из Treeview
            treeview.delete(selected_item)
            # Удаление соответствующей строки из таблицы Excel
            index = int(selected_item[1:]) - 1
            sheet.delete_rows(index + 2)
            # Сохранение изменений в файле Excel
            workbook.save(path)
    else:
        messagebox.showwarning("Предупреждение", "Выберите элемент для удаления.")

def delete():
    pass

def search(sheet, workbook, path):
    path1 = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    workbook1 = openpyxl.load_workbook(path1)
    sheet = workbook1.active
    print(path1)

    load_data(treeview, sheet)
   # data = []
   # for row in sheet.iter_rows(values_only=True):
   #     data.append(row)
   # top = tk.Toplevel()

  #  tree = ttk.Treeview(top)
  #  columnses = data[0]
  #  tree["columns"] = columnses
    # определяем заголовки
  #  for colum in columnses:
  #      tree.column(colum, width=100, minwidth=100)
  #      tree.heading(colum, text=colum)
  #  for row in data[1:]:
 #       tree.insert("", "end", values=row)
#    tree.pack()

def load_data(treeview, sheet):
    
    treeview.delete(*treeview.get_children())
    
    list_values = []
    for row in sheet.iter_rows(values_only = True):
        list_values.append(row)
        
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)


def insert_row(path, sheet, workbook):
    name = name_entry.get()
    age = int(age_spinbox.get())
    course = course_spinbox.get()
    group = group_entry.get()
    #subscription_status = status_combobox.get()
    #employment_status = "Employed" if a else "Unemployed"

    #print(name, age, course, group)


    
    #row_values = [name, age, status, employment, course, group]
    row_values = [name, age, course, group]
    print(path)
    sheet.append(row_values)
    workbook.save(path)

    treeview.insert('', tk.END, values=row_values)

    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    age_spinbox.delete(0, "end")
    age_spinbox.insert(0, "Age")
    #status_combobox.delete(0, "end")
    #status_comboboxx.insert(0, "Subscription")
    #a.delete(0, "end")
    #a.insert(0, "Employment")
    course_spinbox.delete(0, "end")
    course_spinbox.insert(0, "Course")
    group_entry.delete(0, "end")
    group_entry.insert(0, "Group")




def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")



def students1():
    top.destroy()
    root = tk.Tk()
    root.title("ExcelApp Students")
    iconPhoto = tk.PhotoImage(file="icon.png")
    root.iconphoto(False, iconPhoto)

    #global style
    style = ttk.Style(root)
    root.tk.call("source", "forest-light.tcl")
    root.tk.call("source", "forest-dark.tcl")
    style.theme_use("forest-dark")

    combo_list = ["Subscribed", "Not Subscribed", "Other"]

    frame = ttk.Frame(root)
    frame.pack()

    widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
    widgets_frame.grid(row=0, column=0, padx=20, pady=10)

    name_entry = ttk.Entry(widgets_frame)
    name_entry.insert(0, "Name")
    name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
    name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

    age_spinbox = ttk.Spinbox(widgets_frame, from_=18, to=100)
    age_spinbox.insert(0, "Age")
    age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")

    course_entry = ttk.Entry(widgets_frame)
    course_entry.insert(0, "Course")
    course_entry.bind("<FocusIn>", lambda e: course_entry.delete('0', 'end'))
    course_entry.grid(row=2, column=0, padx=5, pady=5, sticky="ew")

    # a = tk.BooleanVar()
    group_entry = ttk.Entry(widgets_frame)
    group_entry.insert(0, "Group")
    group_entry.bind("<FocusIn>", lambda e: group_entry.delete('0', 'end'))
    group_entry.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

    button = ttk.Button(widgets_frame, text="Insert", command=insert_row)
    button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")




    # 1
    button = ttk.Button(widgets_frame, text="Delete", command=delete)
    button.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")


    # 2
    button = ttk.Button(widgets_frame, text="Search", command=search)
    button.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")




    separator = ttk.Separator(widgets_frame)
    separator.grid(row=7, column=0, padx=(20, 10), pady=10, sticky="ew")

    mode_switch = ttk.Checkbutton(
        widgets_frame, text="Mode", style="Switch", command=toggle_mode)
    mode_switch.grid(row=8, column=0, padx=5, pady=10, sticky="nsew")

    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")

    cols = ("Name", "Age", "Course", "Group")
    treeview = ttk.Treeview(treeFrame, show="headings",
                            yscrollcommand=treeScroll.set, columns=cols, height=15)
    treeview.column("Name", width=100)
    treeview.column("Age", width=50)
    treeview.column("Course", width=100) # Subscription
    treeview.column("Group", width=100) # Employment
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data(treeview)


    root.mainloop()

def students():
    top.destroy()
    root = tk.Tk()
    root.title("ExcelApp Students")
    iconPhoto = tk.PhotoImage(file="icon.png")
    root.iconphoto(False, iconPhoto)

    #global path
    #global workbook
    #global sheet
    path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    print(path)

    global style
    style = ttk.Style(root)
    root.tk.call("source", "forest-light.tcl")
    root.tk.call("source", "forest-dark.tcl")
    style.theme_use("forest-dark")

    combo_list = ["Subscribedde", "Not Subscribedde", "Othered"]

    frame = ttk.Frame(root)
    frame.pack()

    widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
    widgets_frame.grid(row=0, column=0, padx=20, pady=10)

    global name_entry, age_spinbox, status_combobox, a, course_spinbox, group_entry
    name_entry = ttk.Entry(widgets_frame)
    name_entry.insert(0, "Name")
    name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
    name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

    age_spinbox = ttk.Spinbox(widgets_frame, from_=18, to=100)
    age_spinbox.insert(0, "Age")
    age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
    
    #status_combobox = ttk.Combobox(widgets_frame, values=combo_list)
    #status_combobox.current(0)
    #status_combobox.grid(row=2, column=0, padx=5, pady=5,  sticky="ew")

    #a = tk.BooleanVar()
    #checkbutton = ttk.Checkbutton(widgets_frame, text="Employed", variable=a)
    #checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

    course_spinbox = ttk.Spinbox(widgets_frame, from_=1, to=5)
    course_spinbox.insert(0, "Course")
    course_spinbox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
    
    #course_entry = ttk.Entry(widgets_frame)
    #course_entry.insert(0, "Stage")
    #course_entry.bind("<FocusIn>", lambda e: course_entry.delete('0', 'end'))
    #course_entry.grid(row=4, column=0, padx=5, pady=5, sticky="ew")

    # a = tk.BooleanVar()
    group_entry = ttk.Entry(widgets_frame)
    group_entry.insert(0, "Group")
    group_entry.bind("<FocusIn>", lambda e: group_entry.delete('0', 'end'))
    group_entry.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

    




    


    # 2
    button3 = ttk.Button(widgets_frame, text="Search", command=lambda: search(sheet, workbook, path))
    button3.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")




    separator = ttk.Separator(widgets_frame)
    separator.grid(row=7, column=0, padx=(20, 10), pady=10, sticky="ew")

    global mode_switch
    mode_switch = ttk.Checkbutton(
        widgets_frame, text="Mode", style="Switch", command=toggle_mode)
    mode_switch.grid(row=8, column=0, padx=5, pady=10, sticky="nsew")

    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")

    #cols = ("Name", "Age", "Subscription", "Employment", "Stage", "Department")
    cols = ("Name", "Age", "Course", "Group")
    global treeview
    treeview = ttk.Treeview(treeFrame, show="headings",
                            yscrollcommand=treeScroll.set, columns=cols, height=15)
    treeview.column("Name", width=100)
    treeview.column("Age", width=50)
    #treeview.column("Subscription", width=100)
    #treeview.column("Employment", width=100)
    treeview.column("Course", width=100) # Subscription
    treeview.column("Group", width=100) # Employment
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data(treeview, sheet)

    button = ttk.Button(widgets_frame, text="Insert", command=lambda: insert_row(path, sheet, workbook))
    button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")
    
    # 1
    button2 = ttk.Button(widgets_frame, text="Delete", command=delete_selected_item)
    button2.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")

    root.mainloop()

def workers():
    top.destroy()
    root = tk.Tk()
    root.title("ExcelApp Workers")
    iconPhoto = tk.PhotoImage(file="icon.png")
    root.iconphoto(False, iconPhoto)

    #global path
    #global workbook
    #global sheet
    path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    print(path)

    global style
    style = ttk.Style(root)
    root.tk.call("source", "forest-light.tcl")
    root.tk.call("source", "forest-dark.tcl")
    style.theme_use("forest-dark")

    combo_list = ["Subscribedde", "Not Subscribedde", "Othered"]

    frame = ttk.Frame(root)
    frame.pack()

    widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
    widgets_frame.grid(row=0, column=0, padx=20, pady=10)

    global name_entry, age_spinbox, status_combobox, a, course_spinbox, group_entry
    name_entry = ttk.Entry(widgets_frame)
    name_entry.insert(0, "Name")
    name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
    name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

    age_spinbox = ttk.Spinbox(widgets_frame, from_=18, to=100)
    age_spinbox.insert(0, "Age")
    age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
    
    #status_combobox = ttk.Combobox(widgets_frame, values=combo_list)
    #status_combobox.current(0)
    #status_combobox.grid(row=2, column=0, padx=5, pady=5,  sticky="ew")

    #a = tk.BooleanVar()
    #checkbutton = ttk.Checkbutton(widgets_frame, text="Employed", variable=a)
    #checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

    course_spinbox = ttk.Spinbox(widgets_frame, from_=1, to=70)
    course_spinbox.insert(0, "Stage")
    course_spinbox.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
    
    #course_entry = ttk.Entry(widgets_frame)
    #course_entry.insert(0, "Stage")
    #course_entry.bind("<FocusIn>", lambda e: course_entry.delete('0', 'end'))
    #course_entry.grid(row=4, column=0, padx=5, pady=5, sticky="ew")

    # a = tk.BooleanVar()
    group_entry = ttk.Entry(widgets_frame)
    group_entry.insert(0, "Department")
    group_entry.bind("<FocusIn>", lambda e: group_entry.delete('0', 'end'))
    group_entry.grid(row=3, column=0, padx=5, pady=5, sticky="ew")

    




    


    # 2
    button3 = ttk.Button(widgets_frame, text="Search", command=lambda: search(sheet))
    button3.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")




    separator = ttk.Separator(widgets_frame)
    separator.grid(row=7, column=0, padx=(20, 10), pady=10, sticky="ew")

    global mode_switch
    mode_switch = ttk.Checkbutton(
        widgets_frame, text="Mode", style="Switch", command=toggle_mode)
    mode_switch.grid(row=8, column=0, padx=5, pady=10, sticky="nsew")

    

    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")

    #cols = ("Name", "Age", "Subscription", "Employment", "Stage", "Department")
    cols = ("Name", "Age", "Stage", "Department")
    global treeview
    treeview = ttk.Treeview(treeFrame, show="headings",
                            yscrollcommand=treeScroll.set, columns=cols, height=15)
    treeview.column("Name", width=100)
    treeview.column("Age", width=50)
    #treeview.column("Subscription", width=100)
    #treeview.column("Employment", width=100)
    treeview.column("Stage", width=100) # Subscription
    treeview.column("Department", width=100) # Employment
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    load_data(treeview, sheet)

    button = ttk.Button(widgets_frame, text="Insert", command=lambda: insert_row(path, sheet, workbook))
    button.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")
    
    # 1
    button2 = ttk.Button(widgets_frame, text="Delete", command=delete_selected_item)
    button2.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")

    root.mainloop()

global k
k = 1
while(k):
    top = tk.Tk()
    top.geometry("700x300")
    style = ttk.Style(top)
    top.tk.call("source", "forest-light.tcl")
    top.tk.call("source", "forest-dark.tcl")
    style.theme_use("forest-dark")
    label = ttk.Label(top, text = "CHOSE TABLE", font=("Arial", 20))
    label.place(x = 275, y = 100)
    button = ttk.Button(text = "Search student table", command = students)
    button.place(x = 300, y = 180)
    button = ttk.Button(text = "Search worker table", command = workers)
    button.place(x = 303, y = 230)
    top.mainloop()
